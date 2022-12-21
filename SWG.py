import coin
import openpyxl
import market

swag = coin.Coin()
swag.origin_price = 1000
DATE = 0
cell = 2

wb = openpyxl.Workbook()
sheet = wb.active
sheet.cell(row=1, column=2).value = "거래량"
sheet.cell(row=1, column=3).value = "시가"
sheet.cell(row=1, column=4).value = "고가"
sheet.cell(row=1, column=5).value = "저가"
sheet.cell(row=1, column=6).value = "종가"
sheet.cell(row=1, column=7).value = "등락률"

def run():
    global cell
    swag.month_setting()
    DATE = swag.date_range()

    ymstring = str(swag.year) + " - " + str(swag.month)
    sheet.cell(row=cell, column=1).value = ymstring
    sheet.cell(row=cell, column=3).value = round(swag.s_price, 2)

    for i in range(1, DATE):
        swag.UpAndDown()
        swag.day()
        swag.month_MAX_MIN()
    swag.vol = int(swag.volume() * 100)
    swag.year_MAX_MIN()

    sheet.cell(row=cell, column=2).value = swag.vol
    sheet.cell(row=cell, column=4).value = round(swag.month_max, 2)
    sheet.cell(row=cell, column=5).value = round(swag.month_min, 2)
    sheet.cell(row=cell, column=6).value = round(swag.e_price, 2)
    sheet.cell(row=cell, column=7).value = round(swag.fluctation, 2)

    wb.save('SWGrecord.xlsx')  #엑셀에 저장
    cell+=1